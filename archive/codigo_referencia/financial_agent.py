"""
Análise financeira: ROI líquido (lance, comissão de leiloeiro, ITBI, registro, reforma),
status `aprovado` quando ROI > limiar (FINANCIAL_ROI_THRESHOLD_PCT), e exportação Excel.

Schema Supabase: valor_arrematacao, valor_mercado_estimado, valor_venda_sugerido, valor_venda_liquido,
lance_maximo_recomendado, fator_liquidez_venda, custo_reforma_estimado, roi_projetado,
status (pendente | analisado | aprovado | descartado_triagem).
"""

from __future__ import annotations

import json
import logging
import os
import re
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field, field_validator
from supabase import Client
from tenacity import before_sleep_log, retry, stop_after_attempt, wait_exponential

from agno.agent import Agent
from agno.models.openai import OpenAIChat
from agno.tools import tool

from ingestion_agent import SUPABASE_TABLE, get_supabase_client
from leilao_constants import (
    STATUS_ANALISADO,
    STATUS_APROVADO,
    STATUS_DESCARTADO_TRIAGEM,
    STATUS_PENDENTE,
)

load_dotenv()

logger = logging.getLogger(__name__)

# Compat: nome antigo = status atual no BD
STATUS_OPORTUNIDADE_A_PLUS = STATUS_APROVADO


def _env_float(name: str, default: float) -> float:
    raw = os.getenv(name)
    if raw is None or raw.strip() == "":
        return default
    return float(raw.replace(",", "."))


ROI_THRESHOLD_PCT_DEFAULT = _env_float("FINANCIAL_ROI_THRESHOLD_PCT", 40.0)
DEFAULT_COMISSAO_LEILOEIRO_PCT = _env_float("DEFAULT_COMISSAO_LEILOEIRO_PCT", 5.0)
DEFAULT_ITBI_PCT = _env_float("DEFAULT_ITBI_PCT", 3.0)
DEFAULT_CUSTOS_REGISTRO = _env_float("DEFAULT_CUSTOS_REGISTRO", 0.0)
DEFAULT_FATOR_LIQUIDEZ = _env_float("FINANCIAL_FATOR_LIQUIDEZ", 0.92)


class ParametrosFinanceirosGlobais(BaseModel):
    """Defaults reconfiguráveis via .env ou instância."""

    model_config = ConfigDict(str_strip_whitespace=True)

    roi_threshold_pct: float = Field(default=ROI_THRESHOLD_PCT_DEFAULT, gt=0, le=500)
    comissao_leiloeiro_pct: float = Field(default=DEFAULT_COMISSAO_LEILOEIRO_PCT, ge=0, le=100)
    itbi_pct: float = Field(default=DEFAULT_ITBI_PCT, ge=0, le=100)
    custos_registro: float = Field(default=DEFAULT_CUSTOS_REGISTRO, ge=0)
    fator_liquidez_venda: float = Field(
        default=DEFAULT_FATOR_LIQUIDEZ,
        gt=0,
        le=1.5,
        description="Margem de segurança aplicada sobre valor de venda estimado",
    )


class RoiCalculoEntrada(BaseModel):
    """Entradas para o ROI; comissão zera em venda direta Caixa."""

    model_config = ConfigDict(str_strip_whitespace=True)

    valor_lance: float = Field(
        ...,
        gt=0,
        description="Lance nominal (arremate); comissão do leiloeiro e ITBI s/ lance incidem sobre este valor",
    )
    valor_venda_estimado: float = Field(..., gt=0, description="Valor de mercado estimado (venda)")
    custo_reforma: float = Field(default=0.0, ge=0)
    comissao_leiloeiro_pct: float = Field(default=DEFAULT_COMISSAO_LEILOEIRO_PCT, ge=0, le=100)
    itbi_pct: float = Field(default=DEFAULT_ITBI_PCT, ge=0, le=100)
    custos_registro: float = Field(default=DEFAULT_CUSTOS_REGISTRO, ge=0)
    fator_liquidez_venda: float = Field(default=DEFAULT_FATOR_LIQUIDEZ, gt=0, le=1.5)
    desconto_avista_pct: float = Field(
        default=0.0,
        ge=0,
        le=99,
        description="Percentual descontado do caixa do lance; comissão e ITBI s/ lance seguem sobre o nominal",
    )
    venda_direta_caixa: bool = Field(
        default=False,
        description="Se True, não incide comissão de leiloeiro",
    )

    @field_validator("comissao_leiloeiro_pct", mode="before")
    @classmethod
    def none_to_default(cls, v: Any) -> Any:
        return DEFAULT_COMISSAO_LEILOEIRO_PCT if v is None else v


class RoiCalculoResultado(BaseModel):
    """Detalhamento do investimento e ROI líquido."""

    valor_lance: float
    valor_lance_efetivo: float = Field(
        ge=0,
        description="Caixa pago pelo lance após desconto à vista (nominal × (1 − desconto%/100))",
    )
    desconto_avista_pct_aplicado: float = Field(default=0.0, ge=0, le=99)
    comissao_leiloeiro_valor: float = Field(ge=0)
    itbi_valor: float = Field(ge=0)
    custos_registro: float = Field(ge=0)
    custo_reforma: float = Field(ge=0)
    investimento_total: float = Field(gt=0)
    valor_venda_estimado: float
    valor_venda_liquido: float
    roi_liquido_pct: float  # mesmo valor persistido em `roi_projetado` no Supabase
    venda_direta_caixa: bool
    comissao_leiloeiro_pct_efetivo: float = Field(ge=0, le=100)
    itbi_pct_aplicado: float


def parse_moeda_br_para_float(valor: Any) -> Optional[float]:
    """Converte 'R$ 1.234,56', número ou string numérica em float."""
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    s = re.sub(r"R\$\s*", "", s, flags=re.IGNORECASE).strip()
    s = s.replace(".", "").replace(",", ".") if "," in s and "." in s else s.replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s == ".":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _coalesce_float(row: dict[str, Any], *keys: str) -> Optional[float]:
    for k in keys:
        if k in row and row[k] is not None:
            v = row[k]
            if isinstance(v, (int, float)):
                return float(v)
            p = parse_moeda_br_para_float(v)
            if p is not None:
                return p
    return None


def _coalesce_pct(row: dict[str, Any], key: str, default: float) -> float:
    if key not in row or row[key] is None:
        return default
    v = row[key]
    if isinstance(v, (int, float)):
        return float(v)
    p = parse_moeda_br_para_float(v)
    return default if p is None else p


def _coalesce_bool(row: dict[str, Any], *keys: str) -> bool:
    for k in keys:
        if k not in row or row[k] is None:
            continue
        v = row[k]
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            return bool(v)
        s = str(v).strip().lower()
        if s in ("1", "true", "sim", "s", "yes", "caixa", "venda_direta"):
            return True
    return False


def calcular_roi_liquido(entrada: RoiCalculoEntrada) -> RoiCalculoResultado:
    """
    ROI líquido (%) = ((venda líquida - investimento total) / investimento total) * 100,
    com venda líquida = valor_venda_estimado * fator_liquidez_venda.

    ``valor_lance`` é o lance nominal. Com desconto à vista, o caixa do lance é
    nominal × (1 − desconto_avista_pct/100). Comissão do leiloeiro (se não Caixa) e ITBI s/ lance
    usam o nominal. ``custos_registro`` é informado pelo chamador (no simulador pode incluir ITBI
    sobre venda quando ``itbi_pct`` for zerado).
    """
    lance_nom = float(entrada.valor_lance)
    d_pct = max(0.0, min(float(entrada.desconto_avista_pct or 0.0), 99.0))
    f_desc = 1.0 - (d_pct / 100.0)
    lance_efetivo = lance_nom * f_desc
    pct_com = 0.0 if entrada.venda_direta_caixa else entrada.comissao_leiloeiro_pct
    comissao = lance_nom * (pct_com / 100.0)
    itbi = lance_nom * (entrada.itbi_pct / 100.0)
    registro = entrada.custos_registro
    reforma = entrada.custo_reforma

    investimento = lance_efetivo + comissao + itbi + registro + reforma
    if investimento <= 0:
        raise ValueError("investimento_total deve ser > 0")

    venda_liquida = entrada.valor_venda_estimado * float(entrada.fator_liquidez_venda)
    roi = ((venda_liquida - investimento) / investimento) * 100.0

    return RoiCalculoResultado(
        valor_lance=lance_nom,
        valor_lance_efetivo=round(lance_efetivo, 2),
        desconto_avista_pct_aplicado=d_pct,
        comissao_leiloeiro_valor=round(comissao, 2),
        itbi_valor=round(itbi, 2),
        custos_registro=round(registro, 2),
        custo_reforma=round(reforma, 2),
        investimento_total=round(investimento, 2),
        valor_venda_estimado=entrada.valor_venda_estimado,
        valor_venda_liquido=round(venda_liquida, 2),
        roi_liquido_pct=round(roi, 4),
        venda_direta_caixa=entrada.venda_direta_caixa,
        comissao_leiloeiro_pct_efetivo=pct_com,
        itbi_pct_aplicado=entrada.itbi_pct,
    )


def calcular_lance_maximo_para_roi(
    *,
    valor_venda_estimado: float,
    roi_objetivo_pct: float,
    custo_reforma: float = 0.0,
    comissao_leiloeiro_pct: float = DEFAULT_COMISSAO_LEILOEIRO_PCT,
    itbi_pct: float = DEFAULT_ITBI_PCT,
    custos_registro: float = DEFAULT_CUSTOS_REGISTRO,
    fator_liquidez_venda: float = DEFAULT_FATOR_LIQUIDEZ,
    venda_direta_caixa: bool = False,
    desconto_avista_pct: float = 0.0,
) -> float:
    """
    Resolve lance **nominal** máximo para bater ROI alvo:
    ROI = (venda_liquida - investimento_total) / investimento_total.

    Com desconto à vista ``d`` (%), investimento inclui lance nominal × (1−d/100) no caixa do lance,
    mais comissão e ITBI s/ lance sobre o nominal (comissão zero se Caixa).
    """
    if valor_venda_estimado <= 0:
        raise ValueError("valor_venda_estimado deve ser > 0")
    if roi_objetivo_pct <= -99.0:
        raise ValueError("roi_objetivo_pct inválido")
    venda_liquida = float(valor_venda_estimado) * float(fator_liquidez_venda)
    investimento_max = venda_liquida / (1.0 + (float(roi_objetivo_pct) / 100.0))
    pct_com = 0.0 if venda_direta_caixa else float(comissao_leiloeiro_pct)
    d = max(0.0, min(float(desconto_avista_pct or 0.0), 99.0)) / 100.0
    coef_lance_nominal = (1.0 - d) + (pct_com / 100.0) + (float(itbi_pct) / 100.0)
    if coef_lance_nominal <= 0:
        raise ValueError("Parâmetros de custos inválidos")
    lance_max = (investimento_max - float(custo_reforma) - float(custos_registro)) / coef_lance_nominal
    return round(max(0.0, lance_max), 2)


def montar_entrada_roi_de_registro(
    row: dict[str, Any],
    globais: Optional[ParametrosFinanceirosGlobais] = None,
) -> RoiCalculoEntrada:
    """Monta entrada a partir de um dict do Supabase (snake_case)."""
    g = globais or ParametrosFinanceirosGlobais()

    lance = _coalesce_float(
        row, "valor_arrematacao", "valor_lance", "valor_lance_atual"
    )
    venda = _coalesce_float(
        row,
        "valor_mercado_estimado",
        "valor_venda_sugerido",
        "valor_venda_estimado",
    )
    reforma = _coalesce_float(row, "custo_reforma_estimado", "custo_reforma") or 0.0

    if lance is None or lance <= 0:
        raise ValueError("Registro sem valor_arrematacao (ou lance legado) válido")
    if venda is None or venda <= 0:
        raise ValueError("Registro sem valor de venda estimado (valor_mercado_estimado)")

    comissao = _coalesce_pct(row, "comissao_leiloeiro_pct", g.comissao_leiloeiro_pct)
    itbi = _coalesce_pct(row, "itbi_pct", g.itbi_pct)
    registro = _coalesce_float(row, "custos_registro", "custos_escritura_registro")
    if registro is None:
        registro = g.custos_registro

    venda_caixa = _coalesce_bool(row, "venda_direta_caixa", "venda_direta", "origem_caixa")
    desc_av = _coalesce_pct(row, "desconto_avista_pct", 0.0)
    if desc_av > 99.0:
        desc_av = 99.0

    return RoiCalculoEntrada(
        valor_lance=lance,
        valor_venda_estimado=venda,
        custo_reforma=reforma,
        comissao_leiloeiro_pct=comissao,
        itbi_pct=itbi,
        custos_registro=registro,
        fator_liquidez_venda=g.fator_liquidez_venda,
        venda_direta_caixa=venda_caixa,
        desconto_avista_pct=desc_av,
    )


def _status_apos_roi(roi_pct: float, threshold: float, status_atual: Optional[str]) -> str:
    if roi_pct > threshold:
        return STATUS_APROVADO
    st = (status_atual or "").strip().lower()
    if st == STATUS_APROVADO:
        return STATUS_ANALISADO
    if st in (STATUS_PENDENTE, STATUS_ANALISADO, ""):
        return STATUS_ANALISADO
    return status_atual or STATUS_ANALISADO


@retry(
    stop=stop_after_attempt(4),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    before_sleep=before_sleep_log(logger, logging.WARNING),
    reraise=True,
)
def _supabase_fetch_avaliados_com_venda(client: Client) -> list[dict[str, Any]]:
    resp = (
        client.table(SUPABASE_TABLE)
        .select("*")
        .not_.is_("valor_mercado_estimado", "null")
        .execute()
    )
    return list(getattr(resp, "data", None) or [])


@retry(
    stop=stop_after_attempt(4),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    before_sleep=before_sleep_log(logger, logging.WARNING),
    reraise=True,
)
def _supabase_fetch_a_plus(client: Client) -> list[dict[str, Any]]:
    resp = (
        client.table(SUPABASE_TABLE)
        .select("*")
        .eq("status", STATUS_APROVADO)
        .execute()
    )
    return list(getattr(resp, "data", None) or [])


def _roi_projetado_de_row(row: dict[str, Any]) -> Optional[float]:
    for k in ("roi_projetado", "roi_liquido_pct"):
        v = row.get(k)
        if v is None:
            continue
        try:
            return float(v)
        except (TypeError, ValueError):
            continue
    return None


@retry(
    stop=stop_after_attempt(4),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    before_sleep=before_sleep_log(logger, logging.WARNING),
    reraise=True,
)
def _supabase_update_financeiro(client: Client, imovel_id: str, payload: dict[str, Any]) -> Any:
    return client.table(SUPABASE_TABLE).update(payload).eq("id", imovel_id).execute()


def aplicar_financeiro_a_registro(
    row: dict[str, Any],
    globais: Optional[ParametrosFinanceirosGlobais] = None,
    client: Optional[Client] = None,
    persistir: bool = True,
) -> dict[str, Any]:
    """
    Calcula ROI, define status (A+ se ROI > limiar) e opcionalmente persiste no Supabase.
    Retorna dict com resultado e payload aplicado.
    """
    if (row.get("status") or "").strip().lower() == STATUS_DESCARTADO_TRIAGEM:
        return {
            "id": row.get("id"),
            "pulado": True,
            "motivo": "descartado_triagem",
            "persistido": False,
        }
    g = globais or ParametrosFinanceirosGlobais()
    entrada = montar_entrada_roi_de_registro(row, g)
    res = calcular_roi_liquido(entrada)
    status_novo = _status_apos_roi(res.roi_liquido_pct, g.roi_threshold_pct, row.get("status"))
    lance_maximo = calcular_lance_maximo_para_roi(
        valor_venda_estimado=entrada.valor_venda_estimado,
        roi_objetivo_pct=g.roi_threshold_pct,
        custo_reforma=entrada.custo_reforma,
        comissao_leiloeiro_pct=entrada.comissao_leiloeiro_pct,
        itbi_pct=entrada.itbi_pct,
        custos_registro=entrada.custos_registro,
        fator_liquidez_venda=entrada.fator_liquidez_venda,
        venda_direta_caixa=entrada.venda_direta_caixa,
        desconto_avista_pct=entrada.desconto_avista_pct,
    )

    payload = {
        "roi_projetado": res.roi_liquido_pct,
        "valor_venda_liquido": res.valor_venda_liquido,
        "lance_maximo_recomendado": lance_maximo,
        "fator_liquidez_venda": entrada.fator_liquidez_venda,
        "status": status_novo,
    }

    out: dict[str, Any] = {
        "id": row.get("id"),
        "resultado": res.model_dump(),
        "status_novo": status_novo,
        "roi_threshold_pct": g.roi_threshold_pct,
        "fator_liquidez_venda": g.fator_liquidez_venda,
        "lance_maximo_recomendado": lance_maximo,
        "persistido": False,
    }

    if persistir and row.get("id"):
        cli = client or get_supabase_client()
        _supabase_update_financeiro(cli, str(row["id"]), payload)
        out["persistido"] = True
        logger.info(
            "Financeiro id=%s ROI=%s%% status=%s",
            row["id"],
            res.roi_liquido_pct,
            status_novo,
        )

    return out


def processar_financeiro_imoveis_com_venda_estimada(
    globais: Optional[ParametrosFinanceirosGlobais] = None,
    client: Optional[Client] = None,
) -> list[dict[str, Any]]:
    """Percorre imóveis com `valor_mercado_estimado` preenchido e atualiza ROI/status."""
    cli = client or get_supabase_client()
    rows = _supabase_fetch_avaliados_com_venda(cli)
    resultados: list[dict[str, Any]] = []
    for row in rows:
        if (row.get("status") or "").strip().lower() == STATUS_DESCARTADO_TRIAGEM:
            resultados.append(
                {
                    "id": row.get("id"),
                    "pulado": True,
                    "motivo": "descartado_triagem",
                }
            )
            continue
        try:
            r = aplicar_financeiro_a_registro(row, globais=globais, client=cli, persistir=True)
            resultados.append(r)
        except Exception as e:
            logger.warning("Pulando id=%s: %s", row.get("id"), e)
            resultados.append({"id": row.get("id"), "erro": str(e)})
    return resultados


def processar_financeiro_imoveis_por_urls(
    urls: set[str],
    globais: Optional[ParametrosFinanceirosGlobais] = None,
    client: Optional[Client] = None,
) -> list[dict[str, Any]]:
    """Recalcula ROI apenas para registros cujo `url_leilao` está em `urls`."""
    cli = client or get_supabase_client()
    rows = _supabase_fetch_avaliados_com_venda(cli)
    alvo = [r for r in rows if (r.get("url_leilao") or "") in urls]
    resultados: list[dict[str, Any]] = []
    for row in alvo:
        if (row.get("status") or "").strip().lower() == STATUS_DESCARTADO_TRIAGEM:
            resultados.append(
                {"id": row.get("id"), "pulado": True, "motivo": "descartado_triagem"}
            )
            continue
        try:
            r = aplicar_financeiro_a_registro(row, globais=globais, client=cli, persistir=True)
            resultados.append(r)
        except Exception as e:
            logger.warning("Pulando id=%s: %s", row.get("id"), e)
            resultados.append({"id": row.get("id"), "erro": str(e)})
    return resultados


def _aplicar_formatacao_excel_basica(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = len(str(ws.cell(row=1, column=col).value or ""))
        for r in range(2, min(ws.max_row + 1, 500)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                maxlen = max(maxlen, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 55)
    ws.freeze_panes = "A2"
    wb.save(path)


def _enriquecer_colunas_financeiras_derivadas(
    rows: list[dict[str, Any]],
    *,
    globais: Optional[ParametrosFinanceirosGlobais] = None,
) -> list[dict[str, Any]]:
    g = globais or ParametrosFinanceirosGlobais()
    out: list[dict[str, Any]] = []
    for row in rows:
        r = dict(row)
        try:
            entrada = montar_entrada_roi_de_registro(r, g)
            res = calcular_roi_liquido(entrada)
            r["valor_venda_liquido"] = res.valor_venda_liquido
            r["fator_liquidez_venda"] = entrada.fator_liquidez_venda
            r["lance_maximo_recomendado"] = calcular_lance_maximo_para_roi(
                valor_venda_estimado=entrada.valor_venda_estimado,
                roi_objetivo_pct=g.roi_threshold_pct,
                custo_reforma=entrada.custo_reforma,
                comissao_leiloeiro_pct=entrada.comissao_leiloeiro_pct,
                itbi_pct=entrada.itbi_pct,
                custos_registro=entrada.custos_registro,
                fator_liquidez_venda=entrada.fator_liquidez_venda,
                venda_direta_caixa=entrada.venda_direta_caixa,
                desconto_avista_pct=entrada.desconto_avista_pct,
            )
        except Exception:
            pass
        out.append(r)
    return out


def exportar_imoveis_roi_minimo_para_excel(
    caminho_saida: str | Path,
    roi_minimo_pct: float,
    client: Optional[Client] = None,
) -> Path:
    """
    Exporta imóveis com `roi_projetado` >= `roi_minimo_pct` (após rodar o batch financeiro).
    Ignora `descartado_triagem`. Formatação igual à exportação de aprovados.
    """
    cli = client or get_supabase_client()
    rows = _supabase_fetch_avaliados_com_venda(cli)
    filtrados: list[dict[str, Any]] = []
    for row in rows:
        if (row.get("status") or "").strip().lower() == STATUS_DESCARTADO_TRIAGEM:
            continue
        roi = _roi_projetado_de_row(row)
        if roi is None:
            continue
        if roi >= float(roi_minimo_pct):
            filtrados.append(row)

    path = Path(caminho_saida).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    if not filtrados:
        df = pd.DataFrame(columns=["mensagem"])
        df.loc[0] = [f"Nenhum imóvel com ROI >= {roi_minimo_pct}%"]
        df.to_excel(path, index=False, engine="openpyxl")
        _aplicar_formatacao_excel_basica(path)
        return path

    filtrados = _enriquecer_colunas_financeiras_derivadas(filtrados)
    df = pd.json_normalize(filtrados)
    preferred = [
        "status",
        "roi_projetado",
        "valor_mercado_estimado",
        "valor_venda_liquido",
        "lance_maximo_recomendado",
        "fator_liquidez_venda",
        "valor_venda_sugerido",
        "valor_arrematacao",
        "data_leilao",
        "custo_reforma_estimado",
        "endereco",
        "cidade",
        "estado",
        "bairro",
        "area_util",
        "area_total",
        "quartos",
        "vagas",
        "url_leilao",
        "id",
    ]
    cols = [c for c in preferred if c in df.columns]
    rest = [c for c in df.columns if c not in cols]
    df = df[cols + rest]
    df.to_excel(path, index=False, engine="openpyxl")
    _aplicar_formatacao_excel_basica(path)
    logger.info("Exportados %d imóveis com ROI >= %.2f%% -> %s", len(filtrados), roi_minimo_pct, path)
    return path


def exportar_oportunidades_a_plus_para_excel(
    caminho_saida: str | Path,
    client: Optional[Client] = None,
) -> Path:
    """
    Exporta apenas imóveis com status `aprovado` (ROI acima do limiar) para Excel.
    """
    cli = client or get_supabase_client()
    rows = _supabase_fetch_a_plus(cli)
    path = Path(caminho_saida).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    if not rows:
        df = pd.DataFrame(columns=["mensagem"])
        df.loc[0] = ["Nenhum imóvel com status aprovado"]
        df.to_excel(path, index=False, engine="openpyxl")
        logger.info("Exportação vazia: %s", path)
        return path

    rows = _enriquecer_colunas_financeiras_derivadas(rows)
    df = pd.json_normalize(rows)
    preferred = [
        "status",
        "roi_projetado",
        "valor_mercado_estimado",
        "valor_venda_liquido",
        "lance_maximo_recomendado",
        "fator_liquidez_venda",
        "valor_venda_sugerido",
        "valor_arrematacao",
        "data_leilao",
        "custo_reforma_estimado",
        "endereco",
        "cidade",
        "estado",
        "bairro",
        "area_util",
        "area_total",
        "quartos",
        "vagas",
        "url_leilao",
        "id",
    ]
    cols = [c for c in preferred if c in df.columns]
    rest = [c for c in df.columns if c not in cols]
    df = df[cols + rest]

    df.to_excel(path, index=False, engine="openpyxl")
    _aplicar_formatacao_excel_basica(path)
    logger.info("Exportados %d imóveis aprovados -> %s", len(rows), path)
    return path


# --- Agno (opcional) ---------------------------------------------------------------------------


@tool(show_result=True)
def calcular_roi_liquido_json(payload_json: str) -> str:
    """
    Calcula ROI a partir de JSON: valor_lance (nominal), valor_venda_estimado, custo_reforma (opcional),
    comissao_leiloeiro_pct, itbi_pct, custos_registro, desconto_avista_pct (opcional, 0–99),
    venda_direta_caixa (bool).
    """
    try:
        data = json.loads(payload_json)
        ent = RoiCalculoEntrada.model_validate(data)
        res = calcular_roi_liquido(ent)
        return res.model_dump_json(ensure_ascii=False)
    except Exception as e:
        logger.exception("ROI cálculo falhou")
        return json.dumps({"erro": str(e)}, ensure_ascii=False)


@tool(show_result=True)
def calcular_lance_maximo_json(payload_json: str) -> str:
    """
    Calcula lance **nominal** máximo recomendado para um ROI alvo.
    JSON esperado: valor_venda_estimado, roi_objetivo_pct e custos opcionais; desconto_avista_pct opcional (0–99).
    """
    try:
        data = json.loads(payload_json)
        out = {
            "lance_maximo_recomendado": calcular_lance_maximo_para_roi(
                valor_venda_estimado=float(data["valor_venda_estimado"]),
                roi_objetivo_pct=float(data["roi_objetivo_pct"]),
                custo_reforma=float(data.get("custo_reforma", 0.0)),
                comissao_leiloeiro_pct=float(data.get("comissao_leiloeiro_pct", DEFAULT_COMISSAO_LEILOEIRO_PCT)),
                itbi_pct=float(data.get("itbi_pct", DEFAULT_ITBI_PCT)),
                custos_registro=float(data.get("custos_registro", DEFAULT_CUSTOS_REGISTRO)),
                fator_liquidez_venda=float(data.get("fator_liquidez_venda", DEFAULT_FATOR_LIQUIDEZ)),
                venda_direta_caixa=bool(data.get("venda_direta_caixa", False)),
                desconto_avista_pct=float(data.get("desconto_avista_pct", 0.0)),
            )
        }
        return json.dumps(out, ensure_ascii=False)
    except Exception as e:
        logger.exception("Lance máximo cálculo falhou")
        return json.dumps({"erro": str(e)}, ensure_ascii=False)


@tool(show_result=True)
def processar_financeiro_supabase() -> str:
    """Recalcula ROI e status (A+ se acima do limiar) para todos com valor_mercado_estimado."""
    try:
        res = processar_financeiro_imoveis_com_venda_estimada()
        return json.dumps(res, ensure_ascii=False, default=str)
    except Exception as e:
        logger.exception("Batch financeiro falhou")
        return json.dumps({"erro": str(e)}, ensure_ascii=False)


@tool(show_result=True)
def exportar_oportunidades_a_plus_excel_tool(caminho_saida: str) -> str:
    """Gera arquivo .xlsx só com imóveis status aprovado."""
    try:
        p = exportar_oportunidades_a_plus_para_excel(caminho_saida)
        return json.dumps({"ok": True, "arquivo": str(p)}, ensure_ascii=False)
    except Exception as e:
        logger.exception("Export Excel falhou")
        return json.dumps({"ok": False, "erro": str(e)}, ensure_ascii=False)


@tool(show_result=True)
def exportar_roi_minimo_excel_tool(caminho_saida: str, roi_minimo_pct: float) -> str:
    """Exporta imóveis com roi_projetado >= roi_minimo_pct (objetivo customizado)."""
    try:
        p = exportar_imoveis_roi_minimo_para_excel(caminho_saida, roi_minimo_pct)
        return json.dumps({"ok": True, "arquivo": str(p)}, ensure_ascii=False)
    except Exception as e:
        logger.exception("Export ROI mínimo falhou")
        return json.dumps({"ok": False, "erro": str(e)}, ensure_ascii=False)


def create_financial_agent(
    *,
    model_id: Optional[str] = None,
    markdown: bool = True,
) -> Agent:
    return Agent(
        model=OpenAIChat(id=model_id or os.getenv("OPENAI_CHAT_MODEL", "gpt-4o-mini")),
        tools=[
            calcular_roi_liquido_json,
            calcular_lance_maximo_json,
            processar_financeiro_supabase,
            exportar_oportunidades_a_plus_excel_tool,
            exportar_roi_minimo_excel_tool,
        ],
        instructions=(
            "Você apoia análise financeira de leilões. ROI líquido usa: "
            "Investimento = Lance + comissão (0 se venda_direta_caixa) + ITBI% sobre lance + "
            "custos de registro + reforma. Aplica fator de liquidez sobre a venda estimada antes do ROI. "
            "ROI% = ((venda líquida - investimento) / investimento)*100. "
            "Limiar para status aprovado vem de FINANCIAL_ROI_THRESHOLD_PCT (padrão 40). ROI gravado em roi_projetado. "
            "Use calcular_lance_maximo_json para retornar lance máximo recomendado para ROI alvo. "
            "Use processar_financeiro_supabase para atualizar o banco; exportar_oportunidades_a_plus_excel_tool para aprovados; "
            "exportar_roi_minimo_excel_tool para filtrar por ROI mínimo customizado."
        ),
        markdown=markdown,
    )


__all__ = [
    "ParametrosFinanceirosGlobais",
    "RoiCalculoEntrada",
    "RoiCalculoResultado",
    "STATUS_ANALISADO",
    "STATUS_APROVADO",
    "STATUS_DESCARTADO_TRIAGEM",
    "STATUS_OPORTUNIDADE_A_PLUS",
    "STATUS_PENDENTE",
    "exportar_imoveis_roi_minimo_para_excel",
    "exportar_roi_minimo_excel_tool",
    "aplicar_financeiro_a_registro",
    "calcular_roi_liquido",
    "calcular_roi_liquido_json",
    "calcular_lance_maximo_para_roi",
    "calcular_lance_maximo_json",
    "create_financial_agent",
    "exportar_oportunidades_a_plus_para_excel",
    "montar_entrada_roi_de_registro",
    "parse_moeda_br_para_float",
    "processar_financeiro_imoveis_com_venda_estimada",
    "processar_financeiro_imoveis_por_urls",
    "processar_financeiro_supabase",
    "exportar_oportunidades_a_plus_excel_tool",
]
